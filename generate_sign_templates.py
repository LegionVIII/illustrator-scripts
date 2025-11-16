#!/usr/bin/env python3
"""
Generate Illustrator JSX Script for Sign Templates
Reads sign schedule Excel file and creates JSX to generate template .ai files
"""

import openpyxl
import os
from pathlib import Path

# Sign type dimensions (inches)
SIGN_DIMENSIONS = {
    'A': {'width': 8.0, 'height': 9.5},
    'B': {'width': 8.0, 'height': 9.5},
    'C': {'width': 8.0, 'height': 9.5},
    'E': {'width': 10.0, 'height': 8.0}
}

def read_sign_schedule(excel_path):
    """Read the sign schedule Excel file and extract unique sign types"""
    workbook = openpyxl.load_workbook(excel_path)
    
    # Check if 'ADA Schedule' sheet exists
    if 'ADA Schedule' not in workbook.sheetnames:
        raise ValueError("'ADA Schedule' sheet not found in Excel file")
    
    sheet = workbook['ADA Schedule']
    
    # Find unique sign types (column E - SIGN TYPE)
    sign_types = set()
    
    # Start from row 3 (skip header rows) and read SIGN TYPE column
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
        sign_type = row[4].value  # Column E (index 4) is SIGN TYPE
        if sign_type and isinstance(sign_type, str):
            sign_type = sign_type.strip()
            if sign_type and sign_type in SIGN_DIMENSIONS:
                sign_types.add(sign_type)
    
    return sorted(list(sign_types))

def generate_jsx_script(sign_types, output_path):
    """Generate JSX script to create Illustrator template files"""
    
    jsx_content = '''// Auto-generated JSX Script to Create Sign Template Files
// Generated from sign schedule Excel file

#target illustrator

function main() {
    // Ask user where to save the template files
    var outputFolder = Folder.selectDialog("Select folder to save sign template files");
    
    if (outputFolder === null) {
        alert("No folder selected. Operation cancelled.");
        return;
    }
    
    var createdFiles = [];
    
'''
    
    # Generate code for each sign type
    for sign_type in sign_types:
        width = SIGN_DIMENSIONS[sign_type]['width']
        height = SIGN_DIMENSIONS[sign_type]['height']
        
        jsx_content += f'''
    // ========================================
    // Create Template for Sign Type {sign_type}
    // ========================================
    try {{
        // Create new document - {width}" x {height}"
        var doc{sign_type} = app.documents.add(
            DocumentColorSpace.CMYK,
            {width} * 72,  // width in points
            {height} * 72   // height in points
        );
        
        doc{sign_type}.rulerOrigin = [0, 0];
        
        // Create layers with proper naming for Variables
        var bgLayer = doc{sign_type}.layers.add();
        bgLayer.name = "Background";
        
        var contentLayer = doc{sign_type}.layers.add();
        contentLayer.name = "Content";
        
        var textLayer = doc{sign_type}.layers.add();
        textLayer.name = "Variable Text";
        
        // Set text layer as active
        doc{sign_type}.activeLayer = textLayer;
        
        // Create background rectangle (sign outline)
        doc{sign_type}.activeLayer = bgLayer;
        var bgRect = doc{sign_type}.pathItems.rectangle(
            {height} * 72,      // top
            0,                  // left
            {width} * 72,      // width
            {height} * 72       // height
        );
        bgRect.filled = true;
        bgRect.fillColor = createCMYKColor(0, 0, 0, 10); // Light gray
        bgRect.stroked = true;
        bgRect.strokeColor = createCMYKColor(0, 0, 0, 100); // Black stroke
        bgRect.strokeWidth = 1;
        bgRect.name = "SignOutline";
        
        // Create placeholder text for Room Number
        doc{sign_type}.activeLayer = textLayer;
        var roomNumText = doc{sign_type}.textFrames.add();
        roomNumText.contents = "###";  // Placeholder for room number
        roomNumText.textRange.characterAttributes.size = 48;
        roomNumText.textRange.characterAttributes.fillColor = createCMYKColor(0, 0, 0, 100);
        roomNumText.textRange.paragraphAttributes.justification = Justification.CENTER;
        roomNumText.name = "RoomNumber";
        
        // Position room number at top center
        roomNumText.top = ({height} * 72) - 36;
        roomNumText.left = (({width} * 72) - roomNumText.width) / 2;
        
        // Create placeholder text for Room Name
        var roomNameText = doc{sign_type}.textFrames.add();
        roomNameText.contents = "ROOM NAME";  // Placeholder for room name
        roomNameText.textRange.characterAttributes.size = 24;
        roomNameText.textRange.characterAttributes.fillColor = createCMYKColor(0, 0, 0, 100);
        roomNameText.textRange.paragraphAttributes.justification = Justification.CENTER;
        roomNameText.name = "RoomName";
        
        // Position room name at center
        roomNameText.top = ({height} * 72) / 2;
        roomNameText.left = (({width} * 72) - roomNameText.width) / 2;
        
        // Add a guide note layer
        var noteLayer = doc{sign_type}.layers.add();
        noteLayer.name = "Notes (Delete Before Production)";
        doc{sign_type}.activeLayer = noteLayer;
        
        var noteText = doc{sign_type}.textFrames.add();
        noteText.contents = "Type {sign_type} Template - {width}\\" x {height}\\"\\n\\nVariable Layers:\\n- RoomNumber\\n- RoomName\\n\\nDelete this notes layer before production.";
        noteText.textRange.characterAttributes.size = 10;
        noteText.textRange.characterAttributes.fillColor = createCMYKColor(100, 0, 100, 0); // Magenta
        noteText.name = "TemplateNotes";
        noteText.top = 20;
        noteText.left = 10;
        
        // Save the file
        var saveFile = new File(outputFolder + "/SignType_{sign_type}_Template.ai");
        doc{sign_type}.saveAs(saveFile);
        doc{sign_type}.close();
        
        createdFiles.push("SignType_{sign_type}_Template.ai");
        
    }} catch (e) {{
        alert("Error creating Type {sign_type} template: " + e);
    }}
'''
    
    # Add helper function and completion message
    jsx_content += '''
    // Show completion message
    if (createdFiles.length > 0) {
        alert("Template files created successfully!\\n\\n" + 
              "Files created: " + createdFiles.length + "\\n\\n" +
              createdFiles.join("\\n") + "\\n\\n" +
              "Location: " + outputFolder.fsName);
    } else {
        alert("No template files were created.");
    }
}

// Helper function to create CMYK color
function createCMYKColor(c, m, y, k) {
    var color = new CMYKColor();
    color.cyan = c;
    color.magenta = m;
    color.yellow = y;
    color.black = k;
    return color;
}

// Run the script
main();
'''
    
    # Write JSX file
    with open(output_path, 'w') as f:
        f.write(jsx_content)
    
    return jsx_content

def main():
    """Main execution function"""
    
    # Path to the Excel file
    excel_path = "/Users/sadmin/Documents/2025 EBCO projects/Rockdale ISD Rockdale Career Academy/Project Management - Rockdale CA/Sign Schedule - Rockdale Career Academy.xlsx"
    
    # Output JSX file path
    script_dir = Path(__file__).parent
    output_jsx = script_dir / "create_sign_templates.jsx"
    
    print("=" * 60)
    print("Sign Template Generator")
    print("=" * 60)
    print(f"\nReading sign schedule from:")
    print(f"  {excel_path}\n")
    
    try:
        # Read the Excel file
        sign_types = read_sign_schedule(excel_path)
        
        print(f"Found {len(sign_types)} unique sign types:")
        for st in sign_types:
            dims = SIGN_DIMENSIONS[st]
            print(f"  - Type {st}: {dims['width']}\" × {dims['height']}\"")
        
        # Generate JSX script
        print(f"\nGenerating JSX script...")
        generate_jsx_script(sign_types, output_jsx)
        
        print(f"\n✓ JSX script created successfully!")
        print(f"  Location: {output_jsx}")
        print("\n" + "=" * 60)
        print("NEXT STEPS:")
        print("=" * 60)
        print("1. Open Adobe Illustrator")
        print("2. Go to: File → Scripts → Other Script...")
        print(f"3. Select: {output_jsx}")
        print("4. Choose where to save your template files")
        print("5. The script will create template .ai files for each sign type")
        print("\nThe template files will include:")
        print("  - Correct artboard dimensions")
        print("  - Background/outline layer")
        print("  - Variable text layers (RoomNumber, RoomName)")
        print("  - Placeholder text")
        print("  - Helpful notes layer")
        print("=" * 60)
        
    except FileNotFoundError:
        print(f"\n✗ ERROR: Excel file not found at:")
        print(f"  {excel_path}")
        print("\nPlease verify the file path and try again.")
    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())
